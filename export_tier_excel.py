import json
import re
import urllib.request
import urllib.error
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=== Starting Golden / Green / Red Data Extraction (April to Till Date) ===")

# 1. Parse Supabase credentials
with open('src/supabaseClient.js', 'r', encoding='utf-8') as f:
    content = f.read()
    url_match = re.search(r"supabaseUrl\s*=\s*['\"]([^'\"]+)['\"]", content)
    key_match = re.search(r"supabaseKey\s*=\s*['\"]([^'\"]+)['\"]", content)
    supabase_url = url_match.group(1)
    supabase_key = key_match.group(1)

# 2. Login to Supabase
login_url = f"{supabase_url}/auth/v1/token?grant_type=password"
login_data = json.dumps({
    "email": "manannegi17@gmail.com",
    "password": "Manan@dyno@17"
}).encode('utf-8')

req = urllib.request.Request(
    login_url,
    data=login_data,
    headers={"apikey": supabase_key, "Content-Type": "application/json"},
    method="POST"
)

with urllib.request.urlopen(req) as res:
    login_res = json.loads(res.read().decode('utf-8'))
    access_token = login_res.get("access_token")

print("Successfully authenticated with Supabase.")

# 3. Fetch all uploaded_files rows from Supabase
url = f"{supabase_url}/rest/v1/uploaded_files?select=id,name,upload_date,record_count,data"
fetch_req = urllib.request.Request(
    url,
    headers={
        "apikey": supabase_key,
        "Authorization": f"Bearer {access_token}"
    }
)

print("Fetching files from Supabase database...")
all_files = []
with urllib.request.urlopen(fetch_req) as res:
    all_files = json.loads(res.read().decode('utf-8'))

print(f"Total uploaded file chunks fetched: {len(all_files)}")

# 4. Separate sales, return, and inventory files
sales_rows = []
return_rows = []
inventory_map = {}
inventory_date = None

for file_obj in all_files:
    fname = file_obj.get('name', '')
    fdata = file_obj.get('data', [])
    if isinstance(fdata, str):
        try:
            fdata = json.loads(fdata)
        except Exception:
            fdata = []
    
    if not isinstance(fdata, list):
        continue

    if fname.startswith('[INVENTORY]'):
        # Latest inventory
        for row in fdata:
            sku = row.get('item_color') or row.get('itemcolor') or row.get('barcode') or row.get('SKU')
            qty = row.get('inventory_qty') or row.get('qty') or row.get('Inventory') or 0
            if sku:
                inventory_map[str(sku).strip()] = int(qty)
                if not inventory_date and row.get('date'):
                    inventory_date = str(row.get('date'))
    elif fname.startswith('[RETURN]'):
        for row in fdata:
            # Check FY26 filter (fy == '2026' or default if April-August files)
            fy = row.get('fy') or '2026'
            if fy == '2026' or '2025' not in fname:
                return_rows.append(row)
    else:
        # Sales files
        for row in fdata:
            fy = row.get('fy') or '2026'
            if fy == '2026' and 'FY25' not in fname:
                sales_rows.append(row)

print(f"Total FY26 Sales Rows Processed: {len(sales_rows):,}")
print(f"Total FY26 Return Rows Processed: {len(return_rows):,}")
print(f"Total Unique SKUs in Inventory Map: {len(inventory_map):,}")

# 5. Process Return Counts per SKU
return_sku_map = {}
for r in return_rows:
    sku = str(r.get('item_color') or r.get('itemcolor') or r.get('barcode') or '').strip()
    if sku:
        return_sku_map[sku] = return_sku_map.get(sku, 0) + 1

# 6. Process Sales Aggregation per SKU
sku_data = {}
for s in sales_rows:
    sku = str(s.get('item_color') or s.get('itemcolor') or s.get('barcode') or '').strip()
    if not sku:
        continue
    
    price_val = float(s.get('priceVal') or s.get('revenue') or s.get('price') or 0)
    division = str(s.get('division') or '-').strip()
    category = str(s.get('categories') or s.get('category') or '-').strip()

    if sku not in sku_data:
        sku_data[sku] = {
            'sku': sku,
            'division': division,
            'category': category,
            'units': 0,
            'revenue': 0.0
        }
    
    sku_data[sku]['units'] += 1
    sku_data[sku]['revenue'] += price_val

# Convert to list and sort by units descending
sku_list = list(sku_data.values())
sku_list.sort(key=lambda x: x['units'], reverse=True)

total_units_all = sum(x['units'] for x in sku_list)
total_revenue_all = sum(x['revenue'] for x in sku_list)

print(f"Total Unique SKUs Sold: {len(sku_list):,}")
print(f"Total Units Sold Across All SKUs: {total_units_all:,}")
print(f"Total Revenue: INR {total_revenue_all:,.2f}")

# 7. Calculate Tier (Golden 50%, Green 30%, Red 20%) & Cumulative Contribution
cum_units = 0
golden_count = 0
green_count = 0
red_count = 0

master_rows = []

for item in sku_list:
    sku = item['sku']
    units = item['units']
    revenue = item['revenue']
    returns = return_sku_map.get(sku, 0)
    return_pct = (returns / units * 100) if units > 0 else 0.0
    asp = (revenue / units) if units > 0 else 0.0
    inv = inventory_map.get(sku, 0)
    
    prev_pct = (cum_units / total_units_all * 100) if total_units_all > 0 else 0.0
    cum_units += units
    curr_cum_pct = (cum_units / total_units_all * 100) if total_units_all > 0 else 0.0
    sku_contrib_pct = (units / total_units_all * 100) if total_units_all > 0 else 0.0

    if prev_pct < 50.0:
        tier = 'Golden (Top 50%)'
        tier_code = 'Golden'
        golden_count += 1
    elif prev_pct < 80.0:
        tier = 'Green (Next 30%)'
        tier_code = 'Green'
        green_count += 1
    else:
        tier = 'Red (Remaining 20%)'
        tier_code = 'Red'
        red_count += 1

    master_rows.append({
        'Tier': tier_code,
        'Tier Group': tier,
        'SKU': sku,
        'Division': item['division'],
        'Category': item['category'],
        'Units Sold': units,
        'Return Qty': returns,
        'Return (%)': return_pct / 100.0,
        'Total Revenue (₹)': revenue,
        'Avg Selling Price (₹)': asp,
        'Current Inventory': inv,
        'Volume Contribution (%)': sku_contrib_pct / 100.0,
        'Cumulative Contribution (%)': curr_cum_pct / 100.0
    })

print(f"Tier Breakdown Summary:")
print(f" - Golden SKUs (Top 50% Volume): {golden_count:,} SKUs")
print(f" - Green SKUs (Next 30% Volume): {green_count:,} SKUs")
print(f" - Red SKUs (Remaining 20% Volume): {red_count:,} SKUs")

# 8. Create Excel Workbook using openpyxl
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name='Calibri', size=16, bold=True, color='1E1E2F')
font_subtitle = Font(name='Calibri', size=11, italic=True, color='555555')
font_section = Font(name='Calibri', size=13, bold=True, color='1E1E2F')

font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
fill_header = PatternFill(start_color='1E1E2F', end_color='1E1E2F', fill_type='solid')

fill_golden = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
font_golden_bold = Font(name='Calibri', size=11, bold=True, color='92400E')

fill_green = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
font_green_bold = Font(name='Calibri', size=11, bold=True, color='065F46')

fill_red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
font_red_bold = Font(name='Calibri', size=11, bold=True, color='991B1B')

font_bold = Font(name='Calibri', size=11, bold=True)
fill_total = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

thin_border_side = Side(border_style='thin', color='E5E7EB')
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

df_master = pd.DataFrame(master_rows)

def create_tier_sheet(ws, title, df_subset, is_summary=False):
    ws.views.sheetView[0].showGridLines = True
    
    ws.append([title])
    ws.cell(row=1, column=1).font = font_title
    
    ws.append([f"Data Period: April 2026 to Date | Total Volume: {total_units_all:,} units | Total Revenue: INR {total_revenue_all:,.2f}"])
    ws.cell(row=2, column=1).font = font_subtitle
    ws.append([])
    
    if is_summary:
        headers = ['Tier', 'Contribution Target', 'SKU Count', 'SKU Share (%)', 'Units Sold', 'Volume Share (%)', 'Total Revenue (INR)', 'Revenue Share (%)', 'Return Qty', 'Avg Return Rate (%)']
        ws.append(headers)
        header_row = 4
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if col_idx in [1, 2] else align_right
        
        tier_summaries = []
        for code, label, target in [('Golden', 'Golden (Top 50%)', '50.0%'), ('Green', 'Green (Next 30%)', '30.0%'), ('Red', 'Red (Remaining 20%)', '20.0%')]:
            sub = df_master[df_master['Tier'] == code]
            skus = len(sub)
            units = sub['Units Sold'].sum()
            rev = sub['Total Revenue (₹)'].sum()
            rets = sub['Return Qty'].sum()
            
            sku_share = skus / len(df_master) if len(df_master) > 0 else 0
            vol_share = units / total_units_all if total_units_all > 0 else 0
            rev_share = rev / total_revenue_all if total_revenue_all > 0 else 0
            avg_return = rets / units if units > 0 else 0
            
            tier_summaries.append({
                'Tier': code,
                'Target': target,
                'SKU Count': skus,
                'SKU Share': sku_share,
                'Units Sold': units,
                'Volume Share': vol_share,
                'Total Revenue': rev,
                'Revenue Share': rev_share,
                'Return Qty': rets,
                'Avg Return Rate': avg_return
            })
        
        for idx, ts in enumerate(tier_summaries, 5):
            ws.append([
                ts['Tier'], ts['Target'], ts['SKU Count'], ts['SKU Share'],
                ts['Units Sold'], ts['Volume Share'], ts['Total Revenue'],
                ts['Revenue Share'], ts['Return Qty'], ts['Avg Return Rate']
            ])
            
            code = ts['Tier']
            fill_c = fill_golden if code == 'Golden' else fill_green if code == 'Green' else fill_red
            font_c = font_golden_bold if code == 'Golden' else font_green_bold if code == 'Green' else font_red_bold
            
            ws.cell(row=idx, column=1).font = font_c
            ws.cell(row=idx, column=1).fill = fill_c
            ws.cell(row=idx, column=1).alignment = align_center
            
            ws.cell(row=idx, column=2).alignment = align_center
            ws.cell(row=idx, column=3).number_format = '#,##0'
            ws.cell(row=idx, column=4).number_format = '0.0%'
            ws.cell(row=idx, column=5).number_format = '#,##0'
            ws.cell(row=idx, column=6).number_format = '0.0%'
            ws.cell(row=idx, column=7).number_format = '#,##0'
            ws.cell(row=idx, column=8).number_format = '0.0%'
            ws.cell(row=idx, column=9).number_format = '#,##0'
            ws.cell(row=idx, column=10).number_format = '0.0%'
            
            for col_idx in range(1, 11):
                ws.cell(row=idx, column=col_idx).border = thin_border
        
        tot_row = len(tier_summaries) + 5
        ws.append([
            'TOTAL / OVERALL', '100.0%', len(df_master), 1.0,
            total_units_all, 1.0, total_revenue_all, 1.0,
            df_master['Return Qty'].sum(),
            (df_master['Return Qty'].sum() / total_units_all) if total_units_all > 0 else 0
        ])
        
        for col_idx in range(1, 11):
            c = ws.cell(row=tot_row, column=col_idx)
            c.font = font_bold
            c.fill = fill_total
            c.border = thin_border
        
        ws.cell(row=tot_row, column=1).alignment = align_center
        ws.cell(row=tot_row, column=2).alignment = align_center
        ws.cell(row=tot_row, column=3).number_format = '#,##0'
        ws.cell(row=tot_row, column=4).number_format = '0.0%'
        ws.cell(row=tot_row, column=5).number_format = '#,##0'
        ws.cell(row=tot_row, column=6).number_format = '0.0%'
        ws.cell(row=tot_row, column=7).number_format = '#,##0'
        ws.cell(row=tot_row, column=8).number_format = '0.0%'
        ws.cell(row=tot_row, column=9).number_format = '#,##0'
        ws.cell(row=tot_row, column=10).number_format = '0.0%'
        
    else:
        cols = ['Tier', 'SKU', 'Division', 'Category', 'Units Sold', 'Return Qty', 'Return (%)', 'Total Revenue (INR)', 'Avg Selling Price (INR)', 'Current Inventory', 'Volume Contribution (%)', 'Cumulative Contribution (%)']
        ws.append(cols)
        header_row = 4
        for col_idx, c_name in enumerate(cols, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if col_idx in [1, 2, 3, 4] else align_right
        
        start_data_row = 5
        for r_idx, (_, row) in enumerate(df_subset.iterrows(), start=start_data_row):
            ws.append([
                row['Tier'],
                row['SKU'],
                row['Division'],
                row['Category'],
                row['Units Sold'],
                row['Return Qty'],
                row['Return (%)'],
                row['Total Revenue (₹)'],
                row['Avg Selling Price (₹)'],
                row['Current Inventory'],
                row['Volume Contribution (%)'],
                row['Cumulative Contribution (%)']
            ])
            
            tier_code = row['Tier']
            fill_row = fill_golden if tier_code == 'Golden' else fill_green if tier_code == 'Green' else fill_red
            font_row = font_golden_bold if tier_code == 'Golden' else font_green_bold if tier_code == 'Green' else font_red_bold
            
            cell_tier = ws.cell(row=r_idx, column=1)
            cell_tier.font = font_row
            cell_tier.fill = fill_row
            cell_tier.alignment = align_center
            
            ws.cell(row=r_idx, column=2).alignment = align_left
            ws.cell(row=r_idx, column=3).alignment = align_left
            ws.cell(row=r_idx, column=4).alignment = align_left
            
            ws.cell(row=r_idx, column=5).number_format = '#,##0'
            ws.cell(row=r_idx, column=6).number_format = '#,##0'
            ws.cell(row=r_idx, column=7).number_format = '0.0%'
            ws.cell(row=r_idx, column=8).number_format = '#,##0'
            ws.cell(row=r_idx, column=9).number_format = '#,##0'
            ws.cell(row=r_idx, column=10).number_format = '#,##0'
            ws.cell(row=r_idx, column=11).number_format = '0.00%'
            ws.cell(row=r_idx, column=12).number_format = '0.00%'
            
            for col_idx in range(1, 13):
                ws.cell(row=r_idx, column=col_idx).border = thin_border

        tot_row = len(df_subset) + start_data_row
        sum_units = df_subset['Units Sold'].sum()
        sum_rets = df_subset['Return Qty'].sum()
        sum_rev = df_subset['Total Revenue (₹)'].sum()
        avg_asp = sum_rev / sum_units if sum_units > 0 else 0
        avg_ret_pct = sum_rets / sum_units if sum_units > 0 else 0
        tot_vol_contrib = sum_units / total_units_all if total_units_all > 0 else 0
        
        ws.append([
            'TOTAL', f"{len(df_subset):,} SKUs", '-', '-',
            sum_units, sum_rets, avg_ret_pct, sum_rev,
            avg_asp, df_subset['Current Inventory'].sum(),
            tot_vol_contrib, '-'
        ])
        
        for col_idx in range(1, 13):
            c = ws.cell(row=tot_row, column=col_idx)
            c.font = font_bold
            c.fill = fill_total
            c.border = thin_border
            
        ws.cell(row=tot_row, column=1).alignment = align_center
        ws.cell(row=tot_row, column=2).alignment = align_left
        ws.cell(row=tot_row, column=5).number_format = '#,##0'
        ws.cell(row=tot_row, column=6).number_format = '#,##0'
        ws.cell(row=tot_row, column=7).number_format = '0.0%'
        ws.cell(row=tot_row, column=8).number_format = '#,##0'
        ws.cell(row=tot_row, column=9).number_format = '#,##0'
        ws.cell(row=tot_row, column=10).number_format = '#,##0'
        ws.cell(row=tot_row, column=11).number_format = '0.00%'

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row in [1, 2]:
                continue
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Create Sheets
ws1 = wb.create_sheet(title="Tier Summary")
create_tier_sheet(ws1, "Golden, Green & Red Contribution Tier Summary", df_master, is_summary=True)

ws2 = wb.create_sheet(title="Golden SKUs (Top 50%)")
create_tier_sheet(ws2, "Golden Tier SKUs (Top 50% Volume Contributors)", df_master[df_master['Tier'] == 'Golden'])

ws3 = wb.create_sheet(title="Green SKUs (Next 30%)")
create_tier_sheet(ws3, "Green Tier SKUs (Next 30% Volume Contributors)", df_master[df_master['Tier'] == 'Green'])

ws4 = wb.create_sheet(title="Red SKUs (Remaining 20%)")
create_tier_sheet(ws4, "Red Tier SKUs (Remaining 20% Volume Contributors)", df_master[df_master['Tier'] == 'Red'])

ws5 = wb.create_sheet(title="All SKUs Master")
create_tier_sheet(ws5, "All SKUs Complete Master List (Golden, Green, Red)", df_master)

output_filename = "Golden_Green_Red_April_to_Date.xlsx"
wb.save(output_filename)
print(f"\n=======================================================")
print(f"SUCCESS! Excel file successfully generated:")
print(f"File Path: c:\\Users\\Manann\\Desktop\\DynoDashboard\\{output_filename}")
print(f"=======================================================")

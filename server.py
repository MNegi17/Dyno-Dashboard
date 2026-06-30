import io
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

@app.route('/api/generate_report', methods=['POST'])
def generate_report():
    try:
        payload = request.json
        if not payload:
            return jsonify({"error": "No JSON payload provided"}), 400

        sales_list = payload.get('sales_data', [])
        inventory_map = payload.get('inventory_data', {})
        start_month = payload.get('start_month', 'April')
        end_month = payload.get('end_month', 'May')
        year = payload.get('year', '2026')

        # Clean/Format Month & Year labels
        year_label = f"FY{year[-2:]}" if year else "FY26"

        # Initialize DataFrames
        sales_df = pd.DataFrame(sales_list)
        inv_df = pd.DataFrame(list(inventory_map.items()), columns=['SKU', 'Current Inventory'])
        
        # --- SHEET 1: All Inventory (High to Low) ---
        inv_sheet_df = inv_df.sort_values(by='Current Inventory', ascending=False).copy()
        
        # --- SHEET 2: Top Bestsellers (Filtered by Year and selected consecutive 2 months) ---
        if not sales_df.empty:
            # Aggregate sales for the 2 months
            grouped = sales_df.groupby('item_color').agg(
                units_sold=('item_color', 'count'),
                revenue=('priceVal', 'sum'),
                division=('division', lambda x: x.iloc[0] if len(x) > 0 else 'Unknown'),
                category=('categories', lambda x: x.iloc[0] if len(x) > 0 else 'Unknown')
            ).reset_index()
            
            # Merge with current inventory level
            bestsellers_df = pd.merge(grouped, inv_df, left_on='item_color', right_on='SKU', how='left')
            bestsellers_df['Current Inventory'] = bestsellers_df['Current Inventory'].fillna(0).astype(int)
            
            # Sort by units sold high to low
            bestsellers_df = bestsellers_df.sort_values(by='units_sold', ascending=False)
            
            # Reorder columns
            bestsellers_df = bestsellers_df[['item_color', 'division', 'category', 'units_sold', 'revenue', 'Current Inventory']]
            bestsellers_df.columns = ['SKU', 'Division', 'Category', 'Units Sold', 'Revenue', 'Current Inventory']
        else:
            bestsellers_df = pd.DataFrame(columns=['SKU', 'Division', 'Category', 'Units Sold', 'Revenue', 'Current Inventory'])

        # --- SHEET 3: Alarming Data (High Sale and Low/Mid Inventory) ---
        if not bestsellers_df.empty:
            alarming_list = []
            for _, row in bestsellers_df.iterrows():
                sku = row['SKU']
                div = row['Division']
                cat = row['Category']
                units = int(row['Units Sold'])
                revenue = float(row['Revenue'])
                stock = int(row['Current Inventory'])
                
                # Monthly sales velocity (2-month period)
                velocity = units / 2.0
                
                # Days of cover (DOC)
                if velocity == 0:
                    doc = 999.0
                else:
                    doc = (stock / velocity) * 30.0
                
                # Classify risk status
                if stock == 0 and units > 0:
                    status = "CRITICAL: Out of Stock"
                elif doc <= 15.0:
                    status = f"HIGH RISK: Low Stock Cover ({doc:.1f} days)"
                elif doc <= 45.0:
                    status = f"MEDIUM RISK: Mid Stock Cover ({doc:.1f} days)"
                else:
                    status = "LOW RISK: Safe Stock Cover"
                
                # Target 60-day cover replenishment recommendation
                replenish = max(0, int(round((velocity * 2.0) - stock)))
                
                # Only include alarming SKUs (Critical, High, or Medium Risk) with positive sales
                if units > 0 and status != "LOW RISK: Safe Stock Cover":
                    alarming_list.append({
                        'SKU': sku,
                        'Division': div,
                        'Category': cat,
                        'Units Sold (2 Months)': units,
                        'Monthly Sales Velocity': velocity,
                        'Current Inventory': stock,
                        'Days of Cover (DOC)': doc if doc < 999 else "N/A",
                        'Risk Status': status,
                        'Recommended Replenishment Qty': replenish
                    })
            
            alarming_df = pd.DataFrame(alarming_list)
            if not alarming_df.empty:
                alarming_df = alarming_df.sort_values(by='Units Sold (2 Months)', ascending=False)
            else:
                alarming_df = pd.DataFrame(columns=['SKU', 'Division', 'Category', 'Units Sold (2 Months)', 'Monthly Sales Velocity', 'Current Inventory', 'Days of Cover (DOC)', 'Risk Status', 'Recommended Replenishment Qty'])
        else:
            alarming_df = pd.DataFrame(columns=['SKU', 'Division', 'Category', 'Units Sold (2 Months)', 'Monthly Sales Velocity', 'Current Inventory', 'Days of Cover (DOC)', 'Risk Status', 'Recommended Replenishment Qty'])

        # Write DataFrames to Excel Workbook in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            inv_sheet_df.to_excel(writer, sheet_name='All Inventory', index=False)
            bestsellers_df.to_excel(writer, sheet_name='Top Bestsellers', index=False)
            alarming_df.to_excel(writer, sheet_name='Alarming Inventory', index=False)

        # Re-open workbook from output bytes to apply styles via openpyxl
        output.seek(0)
        wb = openpyxl.load_workbook(output)

        # Style Definitions
        purple_header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        # Risk Fill Fills & Fonts
        critical_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # soft red
        critical_font = Font(name="Calibri", size=11, bold=True, color="B71C1C")
        
        medium_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")   # soft orange
        medium_font = Font(name="Calibri", size=11, color="E65100")

        thin_side = Side(border_style="thin", color="E0E0E0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            
            # Get max rows & cols
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Format headers
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = purple_header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[1].height = 28

            # Style data cells
            for r in range(2, max_row + 1):
                ws.row_dimensions[r].height = 20
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = regular_font
                    cell.border = thin_border
                    
                    # Deduce column name
                    col_header = ws.cell(row=1, column=c).value
                    
                    # Standard Alignments and Formats
                    if col_header in ['SKU', 'Division', 'Category', 'Risk Status']:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif col_header in ['Current Inventory', 'Units Sold', 'Units Sold (2 Months)', 'Current Inventory', 'Recommended Replenishment Qty']:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                    elif col_header in ['Monthly Sales Velocity', 'Days of Cover (DOC)']:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0.0'
                    elif col_header == 'Revenue':
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '₹#,##0.00'

                # Apply conditional row coloring on 'Alarming Inventory' sheet
                if sheet_name == 'Alarming Inventory':
                    # Find Risk Status column index
                    status_col_idx = None
                    for c in range(1, max_col + 1):
                        if ws.cell(row=1, column=c).value == 'Risk Status':
                            status_col_idx = c
                            break
                    
                    if status_col_idx:
                        status_val = ws.cell(row=r, column=status_col_idx).value
                        if status_val:
                            if "CRITICAL" in status_val or "HIGH RISK" in status_val:
                                for c in range(1, max_col + 1):
                                    cell = ws.cell(row=r, column=c)
                                    cell.fill = critical_fill
                                    if c in [1, status_col_idx]:  # Bold SKU & Status
                                        cell.font = critical_font
                            elif "MEDIUM RISK" in status_val:
                                for c in range(1, max_col + 1):
                                    cell = ws.cell(row=r, column=c)
                                    cell.fill = medium_fill
                                    if c in [1, status_col_idx]:
                                        cell.font = medium_font

            # Auto-fit columns
            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)
                max_len = 0
                for r in range(1, max_row + 1):
                    val = ws.cell(row=r, column=col).value
                    if val is not None:
                        # Account for currency symbol width padding
                        val_str = f"₹{str(val)}" if ws.cell(row=1, column=col).value == 'Revenue' and r > 1 else str(val)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Save back to a new BytesIO stream
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        filename = f"Intelli_Inventory_Report_{year_label}_{start_month}_{end_month}.xlsx"
        return send_file(
            final_output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/api/generate_weekly_business_report', methods=['POST'])
def generate_weekly_business_report():
    try:
        payload = request.json
        if not payload:
            return jsonify({"error": "No JSON payload provided"}), 400

        sales_list = payload.get('sales_data', [])
        returns_list = payload.get('returns_data', [])
        month = payload.get('month', 'June')
        year = payload.get('year', '2026')
        year_label = f"FY{year[-2:]}" if year else "FY26"

        # Create DataFrames
        sales_df = pd.DataFrame(sales_list)
        returns_df = pd.DataFrame(returns_list)

        # Dynamic Calendar Week Calculation (Weeks start on Sunday)
        import calendar
        import datetime

        month_name_to_num = {
            'January': 1, 'February': 2, 'March': 3, 'April': 4,
            'May': 5, 'June': 6, 'July': 7, 'August': 8,
            'September': 9, 'October': 10, 'November': 11, 'December': 12
        }
        m_num = month_name_to_num.get(month, 6)
        y_num = int(year)

        num_days = calendar.monthrange(y_num, m_num)[1]

        weeks = []
        current_week_days = []
        week_idx = 1

        for d in range(1, num_days + 1):
            dt = datetime.date(y_num, m_num, d)
            # Sunday starts a new week in standard calendar (weekday() == 6)
            if dt.weekday() == 6 and current_week_days:
                start_day = current_week_days[0]
                end_day = current_week_days[-1]
                start_dt = datetime.date(y_num, m_num, start_day)
                end_dt = datetime.date(y_num, m_num, end_day)
                label = f"Week {week_idx} ({start_dt.strftime('%b %d')} - {end_dt.strftime('%b %d')})"
                weeks.append({
                    'week_num': week_idx,
                    'start_day': start_day,
                    'end_day': end_day,
                    'label': label,
                    'days': current_week_days
                })
                current_week_days = []
                week_idx += 1
            current_week_days.append(d)

        if current_week_days:
            start_day = current_week_days[0]
            end_day = current_week_days[-1]
            start_dt = datetime.date(y_num, m_num, start_day)
            end_dt = datetime.date(y_num, m_num, end_day)
            label = f"Week {week_idx} ({start_dt.strftime('%b %d')} - {end_dt.strftime('%b %d')})"
            weeks.append({
                'week_num': week_idx,
                'start_day': start_day,
                'end_day': end_day,
                'label': label,
                'days': current_week_days
            })

        def find_week_index(day, weeks_list):
            for idx, w in enumerate(weeks_list):
                if day in w['days']:
                    return idx
            return 0

        channel_stats = {}

        unique_channels = set()
        if not sales_df.empty and 'channel_name' in sales_df.columns:
            unique_channels.update(sales_df['channel_name'].dropna().unique())
        if not returns_df.empty and 'channel_name' in returns_df.columns:
            unique_channels.update(returns_df['channel_name'].dropna().unique())

        for channel in unique_channels:
            channel_stats[channel] = {
                'Channel': channel,
                'Monthly Sales Qty': 0, 'Monthly Sales Value': 0.0, 'Monthly Returns Qty': 0
            }
            for w_idx in range(len(weeks)):
                channel_stats[channel][f'W{w_idx+1} Sales Qty'] = 0
                channel_stats[channel][f'W{w_idx+1} Sales Value'] = 0.0
                channel_stats[channel][f'W{w_idx+1} Returns Qty'] = 0

        # Process sales
        if not sales_df.empty:
            sales_df['date_dt'] = pd.to_datetime(sales_df['parsedDate'], errors='coerce')
            sales_df['day'] = sales_df['date_dt'].dt.day
            sales_df = sales_df.dropna(subset=['day'])
            sales_df['day'] = sales_df['day'].astype(int)

            for _, row in sales_df.iterrows():
                ch = row['channel_name']
                day = row['day']
                w_idx = find_week_index(day, weeks)
                val = float(row.get('priceVal', 0) or 0)
                if ch in channel_stats:
                    channel_stats[ch][f'W{w_idx+1} Sales Qty'] += 1
                    channel_stats[ch][f'W{w_idx+1} Sales Value'] += val
                    channel_stats[ch]['Monthly Sales Qty'] += 1
                    channel_stats[ch]['Monthly Sales Value'] += val

        # Process returns
        if not returns_df.empty:
            returns_df['date_dt'] = pd.to_datetime(returns_df['parsedDate'], errors='coerce')
            returns_df['day'] = returns_df['date_dt'].dt.day
            returns_df = returns_df.dropna(subset=['day'])
            returns_df['day'] = returns_df['day'].astype(int)

            for _, row in returns_df.iterrows():
                ch = row['channel_name']
                day = row['day']
                w_idx = find_week_index(day, weeks)
                qty = float(row.get('return_qty', 0) or 0)
                if ch in channel_stats:
                    channel_stats[ch][f'W{w_idx+1} Returns Qty'] += qty
                    channel_stats[ch]['Monthly Returns Qty'] += qty

        records = list(channel_stats.values())
        if not records:
            cols = ['Channel']
            for w_idx in range(len(weeks)):
                cols.extend([
                    f'W{w_idx+1} Sales Qty',
                    f'W{w_idx+1} Sales Value',
                    f'W{w_idx+1} Returns Qty'
                ])
            cols.extend([
                'Monthly Sales Qty',
                'Monthly Sales Value',
                'Monthly Returns Qty'
            ])
            df = pd.DataFrame(columns=cols)
        else:
            df = pd.DataFrame(records)
            df = df.sort_values(by='Monthly Sales Value', ascending=False)

        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weekly Business Report"
        ws.views.sheetView[0].showGridLines = True

        purple_header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_side = Side(border_style="thin", color="E0E0E0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24

        ws.merge_cells('A1:A2')
        ws.cell(row=1, column=1, value="Channel")

        # Dynamically build double-headers
        for idx, w in enumerate(weeks):
            start_col = 2 + idx * 3
            end_col = 4 + idx * 3
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            ws.merge_cells(f"{start_letter}1:{end_letter}1")
            ws.cell(row=1, column=start_col, value=w['label'])

        m_start_col = 2 + len(weeks) * 3
        m_end_col = 4 + len(weeks) * 3
        m_start_letter = get_column_letter(m_start_col)
        m_end_letter = get_column_letter(m_end_col)
        ws.merge_cells(f"{m_start_letter}1:{m_end_letter}1")
        ws.cell(row=1, column=m_start_col, value="Total Monthly Performance")

        total_cols = 4 + len(weeks) * 3
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = purple_header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Row 2 Subheaders
        ws.cell(row=2, column=1).fill = purple_header_fill
        ws.cell(row=2, column=1).font = header_font

        for idx in range(len(weeks) + 1):
            base_col = 2 + idx * 3
            ws.cell(row=2, column=base_col, value="Sales (Qty)")
            ws.cell(row=2, column=base_col+1, value="Sales (Value)")
            ws.cell(row=2, column=base_col+2, value="Returns (Qty)")

        for col in range(1, total_cols + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = purple_header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row = 3
        for _, r_data in df.iterrows():
            ws.row_dimensions[current_row].height = 20
            cell_ch = ws.cell(row=current_row, column=1, value=r_data['Channel'])
            cell_ch.font = regular_font
            cell_ch.alignment = Alignment(horizontal="left", vertical="center")
            cell_ch.border = thin_border

            cols = []
            for w_idx in range(len(weeks)):
                cols.extend([
                    f'W{w_idx+1} Sales Qty',
                    f'W{w_idx+1} Sales Value',
                    f'W{w_idx+1} Returns Qty'
                ])
            cols.extend([
                'Monthly Sales Qty',
                'Monthly Sales Value',
                'Monthly Returns Qty'
            ])

            for idx, col_name in enumerate(cols):
                col_idx = idx + 2
                val = r_data[col_name]
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
                if "Value" in col_name:
                    cell.number_format = '₹#,##0.00'
                else:
                    cell.number_format = '#,##0'
            current_row += 1

        ws.row_dimensions[current_row].height = 22
        cell_total_label = ws.cell(row=current_row, column=1, value="Total")
        cell_total_label.font = bold_font
        cell_total_label.alignment = Alignment(horizontal="left", vertical="center")
        
        double_bottom_border = Border(
            left=thin_side, right=thin_side,
            top=thin_side,
            bottom=Side(border_style="double", color="4A148C")
        )
        cell_total_label.border = double_bottom_border

        for col_idx in range(2, total_cols + 1):
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}3:{col_letter}{current_row-1})"
            cell = ws.cell(row=current_row, column=col_idx, value=formula)
            cell.font = bold_font
            cell.border = double_bottom_border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
            if (col_idx - 3) % 3 == 0:
                cell.number_format = '₹#,##0.00'
            else:
                cell.number_format = '#,##0'

        for col_idx in range(1, total_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for r in range(2, current_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                if val is not None:
                    val_str = str(val)
                    if val_str.startswith('='):
                        val_str = "12,345.00"
                    elif (col_idx - 3) % 3 == 0 and r >= 3:
                        val_str = f"₹{val_str}"
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

        wb.save(output)
        output.seek(0)

        filename = f"Weekly_Business_Report_{year_label}_{month}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5001))
    app.run(host='0.0.0.0', port=port, debug=True)


